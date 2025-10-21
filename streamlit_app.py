# streamlit_excel_cleaner.py
import os
import io
import base64
import pandas as pd
import streamlit as st
from io import BytesIO
import csv
import re
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Make headers case/spacing insensitive and map provider variants
    to a canonical set that the rest of the code expects.
    """
    def norm(s: str) -> str:
        return str(s).strip().lower().replace("_", " ")

    canon = {
        # times/dates (common)
        "pickup time (local)": "Pickup Time (Local)",
        "pickup time (utc)": "Pickup Time (UTC)",
        "request time (local)": "Request Time (Local)",
        "request time (utc)": "Request Time (UTC)",
        "pickup date (local)": "Pickup Date (Local)",
        "pickup date (utc)": "Pickup Date (UTC)",
        "request date (local)": "Request Date (Local)",
        "request date (utc)": "Request Date (UTC)",

        # your new variants (exact strings from your message)
        "pick up date": "Pickup Date (Local)",
        "pick up time": "Pickup Time (Local)",
        "rider full name": "Rider Name",
        "pick up address": "Pickup Address",
        "drop off address": "Drop-off Address",
        "pickup to drop-off miles": "Pickup to Drop-off Miles",
        "rider phone #": "Passenger Number",
        "dispatcher email": "Dispatcher Email",
        "internal code": "Internal Code",
        "transaction amount": "Transaction Amount",
        "transaction type": "Transaction Type",

        # identity / contact (general)
        "rider phone number": "Passenger Number",
        "employee id": "Employee ID",
        "member id": "Member ID",
        "rider id": "Rider ID",

        # status / notes
        "ride status": "Ride Status",
        "concierge internal note": "Internal Note",
        "expense note": "Internal Note",

        # addresses (other common labels)
        "pickup address": "Pickup Address",
        "drop-off address": "Drop-off Address",
        "drop off address": "Drop-off Address",  # just in case
        "distance (miles)": "Pickup to Drop-off Miles",  # useful fallback
    }

    df = df.rename(columns={c: canon.get(norm(c), c) for c in df.columns})

    # If only "Rider Name" exists, split into First/Last for grouping/sorting
    if "Rider Name" in df.columns and ("First Name" not in df.columns or "Last Name" not in df.columns):
        name_series = df["Rider Name"].fillna("").astype(str)
        first_last = name_series.apply(_split_first_last)
        df["First Name"] = first_last.apply(lambda x: x[0])
        df["Last Name"]  = first_last.apply(lambda x: x[1])

    return df

# --- Add near the top (e.g., under the other constants) ---

UBER_DETAIL_COLUMNS = [
    "Trip/Eats ID","Transaction Timestamp (UTC)","Request Date (UTC)","Request Time (UTC)",
    "Request Date (Local)","Request Time (Local)","Request Type","Pickup Date (UTC)",
    "Pickup Time (UTC)","Pickup Date (Local)","Pickup Time (Local)","Drop-off Date (UTC)",
    "Drop-off Time (UTC)","Drop-off Date (Local)","Drop-off Time (Local)",
    "Request Timezone Offset from UTC","First Name","Last Name","Email","Employee ID",
    "Service","City","Distance (mi)","Haversine Distance (mi)","Duration (min)","Pickup Address",
    "Pickup Latitude","Pickup Longitude","Drop-off Address","Drop Off Latitude","Drop Off Longitude",
    "Ride Status","Expense Code","Expense Memo","Invoices","Program","Group","Payment Method",
    "Transaction Type","Fare in Local Currency (excl. Taxes)","Taxes in Local Currency",
    "Tip in Local Currency","Transaction Amount in Local Currency (incl. Taxes)","Local Currency Code",
    "Fare in USD (excl. Taxes)","Taxes in USD","Tip in USD","Transaction Amount in USD (incl. Taxes)",
    "Estimated Service and Technology Fee (incl. Taxes, if any) in USD","Health Dashboard URL",
    "Invoice Number","Driver First Name","Guest First Name","Guest Last Name","Guest Phone Number",
    "Deductions in Local Currency","Member ID","Plan ID","Network Transaction Id","IsGroupOrder",
    "Fulfilment Type","Country","Cancellation type","Membership Savings(Local Currency)",
    "Granular Service Purpose Type"
]

def build_uber_detail_sheet(df_raw: pd.DataFrame) -> pd.DataFrame:
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(columns=UBER_DETAIL_COLUMNS)

    raw = df_raw.copy()

    # 🔧 Ensure unique column names before selecting by list
    raw = raw.loc[:, ~raw.columns.duplicated()]

    for col in UBER_DETAIL_COLUMNS:
        if col not in raw.columns:
            raw[col] = ""

    detail = raw[UBER_DETAIL_COLUMNS].copy()
    for c in detail.columns:
        if pd.api.types.is_object_dtype(detail[c]):
            detail[c] = detail[c].fillna("").astype(str).str.strip()
    return detail

# List of columns to hide (delete)
columns_to_hide = [
    "Ride ID", "Pickup Time (UTC)", "Pickup Timezone offset from UTC", "Pickup Date (UTC)",
    "Drop-off Time (Local)", "Drop-off Time (UTC)", "Drop-off Timezone", "Drop-off Date (Local)", "Drop-off Date (UTC)", "Email",
    "Pickup City", "Pickup State", "Pickup Zip Code", "Requester Name",
    "Drop-off City", "Drop-off State", "Drop-off Zip Code",
    "Request Address", "Request City", "Request State", "Request Zip Code",
    "Destination Address", "Destination City", "Destination State", "Destination Zip Code",
    "Duration (minutes)", "Ride Fare", "Ride Fees", "Ride Discounts", "Ride Tip", "Ride Cost",
    "Business Services Fee", "Transaction Date (UTC)", "Transaction Time (UTC)", "Transaction Currency", "Transaction Outcome",
    "Expense Code", "Expense Note", "Ride Type", "Employee ID", "Custom Tag 1", "Custom Tag 2",
    "Fare Type", "Scheduled Ride Id", "Flex Ride Id", "Flex Ride", "Pickup Latitude", "Pickup Longitude",
    "Drop-off Latitude", "Drop-off Longitude",
    "Trip/Eats ID", "Transaction Timestamp (UTC)", "Request Date (UTC)", "Request Time (UTC)", "Request Date (Local)", "Request Time (Local)",
    "Request Type", "Request Timezone Offset from UTC", "Service", "City", "Haversine Distance (mi)", "Duration (min)", "Drop Off Latitude",
    "Drop Off Longitude", "Expense Code", "Invoices", "Program", "Group", "Payment Method", "Fare in Local Currency (excl. Taxes)", "Taxes in Local Currency",
    "Tip in Local Currency", "Taxes in Local Currency", "Tip in Local Currency",
    "Local Currency Code", "Fare in USD (excl. Taxes)", "Taxes in USD", "Tip in USD", "Transaction Amount in USD (incl. Taxes)", "Estimated Service and Technology Fee (incl. Taxes, if any) in USD",
    "Health Dashboard URL", "Invoice Number", "Driver First Name", "Deductions in Local Currency", "Member ID", "Plan ID", "Network Transaction Id",
    "IsGroupOrder", "Fulfilment Type", "Country", "Cancellation type", "Membership Savings(Local Currency)", "Granular Service Purpose Type"
]

DAILY_MIN_COLS = [
    # date/time we might use downstream
    "Pickup Date (Local)", "Request Date (Local)",
    "Pickup Time (Local)", "Request Time (Local)",
    "Pickup Time (UTC)", "Request Time (UTC)",
    # identity
    "First Name", "Last Name", "Rider Name",
    "Passenger Number", "Member ID", "Employee ID", "Rider ID",
    # extras used by the daily builder
    "Pickup Address", "Drop-off Address",
    "Pickup to Drop-off Miles", "Distance (miles)",
    "Transaction Amount", "Dispatcher Email", "Internal Code", "Transaction Type",
    # status / notes
    "Ride Status", "Internal Note",
]

def _ensure_daily_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Make sure df has all DAILY_MIN_COLS and everything is string-like."""
    if df is None or df.empty:
        return pd.DataFrame(columns=DAILY_MIN_COLS)

    out = df.copy()

    # 💡 Drop duplicate column names before reindexing
    out = out.loc[:, ~out.columns.duplicated()]

    # Add any missing columns as empty strings
    for c in DAILY_MIN_COLS:
        if c not in out.columns:
            out[c] = ""

    # Keep only the columns we care about (preserves order)
    out = out[[c for c in DAILY_MIN_COLS if c in out.columns]]

    # Coerce everything to string to avoid type conflicts
    for c in out.columns:
        out[c] = out[c].astype(object).fillna("").astype(str).str.strip()

    return out

internal_note_values = ["FCC", "FCM", "FCSH", "FCSC", "DTF", "DTFCE", "FCEX9", "FCEX10", "FCE10"]

# Keep this where internal_note_values is defined
GROUP_A_NOTES = {"FCC","FCM","FCSH","FCSC","FCEX9","FCEX09","FCEX10", "FCE10", "FCEX9"}
GROUP_B_NOTES = {"DTF", "DTFCE"}

# Union used to validate internal notes (main-block vs invalid footer)
internal_note_values = sorted(list(GROUP_A_NOTES | GROUP_B_NOTES))


def _collapse_ws_series(s: pd.Series) -> pd.Series:
    return (
        s.astype(object)
         .fillna("")
         .astype(str)
         .str.replace(r"\s+", " ", regex=True)
         .str.strip()
    )

def _split_first_last(full: str) -> tuple[str, str]:
    full = re.sub(r"\s+", " ", str(full).strip())  # collapse internal spaces first
    if not full:
        return "", ""
    parts = full.split(" ", 1)  # first token = first name, rest = last name
    if len(parts) == 1:
        return parts[0], ""
    return parts[0].strip(), parts[1].strip()


def _coalesce_duplicate_columns(df: pd.DataFrame, only: list[str] | None = None) -> pd.DataFrame:
    """
    For any duplicated column names, combine them row-wise by taking the first
    non-empty value across the duplicates, keep a single column with that name,
    and drop the rest. If `only` is provided, only coalesce those names.
    """
    cols_list = list(df.columns)
    dup_names = pd.Series(cols_list)[pd.Series(cols_list).duplicated()].unique().tolist()
    if only is not None:
        dup_names = [n for n in dup_names if n in only]

    for name in dup_names:
        # where the first occurrence currently lives
        first_pos = cols_list.index(name)

        # grab all duplicates of this name
        same_name_cols = [c for c in df.columns if c == name]
        block = df[same_name_cols].astype(object).fillna("").astype(str).applymap(str.strip)

        # row-wise: first non-empty wins
        merged = block.apply(lambda row: next((v for v in row if v != ""), ""), axis=1)

        # drop all duplicates of this name
        df = df.drop(columns=same_name_cols)

        # re-insert a single merged column back at the original first position
        df.insert(first_pos, name, merged)

        # refresh list for next iteration
        cols_list = list(df.columns)

    return df

def _col(df: pd.DataFrame, name: str, default: str = "") -> pd.Series:
    """Return a string Series for column `name`, or a default-filled Series if missing."""
    if name in df.columns:
        return df[name].astype(object).fillna(default).astype(str)
    # preserve index/length
    return pd.Series([default] * len(df), index=df.index, dtype="object")


def _is_canceled_series(status_col: pd.Series) -> pd.Series:
    """
    True for any canceled-like status (including refunds).
    """
    s = status_col.fillna("").astype(str).str.upper()
    squeezed = s.str.replace(r"[\s\-_]", "", regex=True)

    canceled_tokens = {
        "CANCELED", "CANCELLED",
        "DRIVERCANCELED", "DRIVER_CANCELLED",
        "USERCANCELED", "USERCANCELLED",
        "RIDERCANCELED", "RIDERCANCELLED",
        "REFUND",  # ← treat refunds like canceled for totals/✓x
    }

    return squeezed.isin(canceled_tokens) | s.str.contains("CANCEL", na=False)

def detect_header(uploaded_file):
    uploaded_file.seek(0)
    for idx in [0, 4, 5]:
        try:
            df = pd.read_csv(uploaded_file, header=idx, nrows=1)
            if any("trip/eats id" in col.lower() for col in df.columns):
                uploaded_file.seek(0)
                print('I found the headers on index:', idx)
                return idx
        except Exception:
            pass
        uploaded_file.seek(0)
    return None

def clean_file_without_headers(df):
    df = _normalize_headers(df)

    # If guest name columns exist, use them
    name_headers = ["First Name", "Last Name", "Guest First Name", "Guest Last Name"]
    if all(h in df.columns for h in name_headers):
        df = df.drop(columns=["First Name", "Last Name"])
        df = df.rename(columns={"Guest First Name": "First Name", "Guest Last Name": "Last Name"})

    # Common renames
    df = df.rename(columns={
        "Distance (mi)": "Pickup to Drop-off Miles",
        "Transaction Amount in Local Currency (incl. Taxes)": "Transaction Amount",
        "Guest Phone Number": "Passenger Number",
        "Expense Memo": "Internal Note",
    })

    # Combine Email variants if present
    if 'Email' in df.columns and 'Requester Email' in df.columns:
        df['Email Info'] = df['Email'].combine_first(df['Requester Email'])
        df.drop(['Email', 'Requester Email'], axis=1, inplace=True)
    elif 'Email' in df.columns:
        df.rename(columns={"Email": "Email Info"}, inplace=True)
    elif 'Requester Email' in df.columns:
        df.rename(columns={"Requester Email": "Email Info"}, inplace=True)

    # Keep everything we might need later
    desired_columns = [
        # date/time
        "Pickup Date (Local)", "Request Date (Local)",
        "Pickup Time (Local)", "Request Time (Local)", "Pickup Time (UTC)", "Request Time (UTC)",
        # identity
        "First Name", "Last Name", "Rider Name",
        "Passenger Number", "Member ID", "Employee ID", "Rider ID",
        # your requested extras
        "Pickup Address", "Drop-off Address",
        "Pickup to Drop-off Miles", "Transaction Amount",
        "Dispatcher Email", "Internal Code", "Transaction Type",
        # status / notes
        "Ride Status", "Internal Note",
    ]
    present = [c for c in desired_columns if c in df.columns]
    return df[present].copy()

def clean_file(uploaded_file):
    try:
        print("\n📥 File received:", uploaded_file.name)
        is_common_courtesy = False

        if uploaded_file.name.endswith(".csv"):
            # ✅ REPLACE your current preview/if-block with this:
            peek = pd.read_csv(uploaded_file, nrows=8, header=None, dtype=str).fillna("")
            uploaded_file.seek(0)
            has_common_courtesy = peek.apply(
                lambda r: r.astype(str).str.contains("common courtesy", case=False, na=False)
            ).any().any()

            if has_common_courtesy:
                header_row = detect_header(uploaded_file)  # your existing helper
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, header=header_row if header_row is not None else 4)
                is_common_courtesy = True
            else:
                df = pd.read_csv(uploaded_file)

            # keep your existing BOM/whitespace cleanup
            df.columns = df.columns.str.replace('\ufeff', '', regex=False).str.strip()

        else:
            # (xlsx branch unchanged)
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.replace('\ufeff', '', regex=False).str.strip()

        # Normalize headers
        df = _normalize_headers(df)

        # Coalesce duplicates introduced by normalization (especially "Internal Note")
        df = _coalesce_duplicate_columns(df, only=["Internal Note"])

        # Keep RAW before hiding/renaming
        df_raw = df.copy()

        # Drop noisy columns
        custom_columns_to_hide = columns_to_hide.copy()

        # Never drop 'Email' so Uber sheets keep it
        if "Email" in custom_columns_to_hide:
            custom_columns_to_hide.remove("Email")

        # (keeps your special-case too; harmless now)
        if is_common_courtesy and "Email" in custom_columns_to_hide:
            custom_columns_to_hide.remove("Email")

        df = df.drop(columns=[c for c in custom_columns_to_hide if c in df.columns], errors="ignore")

        # Now define renames and apply them (do NOT use column_rename_map before this point)
        column_rename_map = {
            "Distance (mi)": "Distance (miles)",
            "Transaction Amount in Local Currency (incl. Taxes)": "Transaction Amount",
            "Guest Phone Number": "Passenger Number",
            "Expense Memo": "Internal Note",
            "Email": "Email Info",
            "Requester Email": "Email Info",
        }
        df.rename(columns=column_rename_map, inplace=True)

        # Ensure Dispatcher Email is populated for Uber (and others)
        if "Dispatcher Email" not in df.columns:
            if "Email Info" in df.columns:
                df["Dispatcher Email"] = df["Email Info"]
            elif "Email" in df.columns:  # fallback if renames change later
                df["Dispatcher Email"] = df["Email"]
            elif "Requester Email" in df.columns:
                df["Dispatcher Email"] = df["Requester Email"]


        # Make names unique in the cleaned view
        df = df.loc[:, ~df.columns.duplicated()]

        # Handle guest name columns if present
        name_headers = ["First Name", "Last Name", "Guest First Name", "Guest Last Name"]
        if all(h in df.columns for h in name_headers):
            df = df.drop(columns=["First Name", "Last Name"])
            df = df.rename(columns={"Guest First Name": "First Name", "Guest Last Name": "Last Name"})

        # Trim basics (NaN-safe)
        for c in ["First Name", "Last Name", "Passenger Number"]:
            if c in df.columns:
                df[c] = df[c].fillna("").astype(str).str.strip()

        # Collapse internal spaces so "J  HARRINGTON" == "J HARRINGTON"
        for c in ["First Name", "Last Name", "Rider Name"]:
            if c in df.columns:
                df[c] = _collapse_ws_series(df[c])


        # 🔧 Normalize phone now so downstream is clean
        if "Passenger Number" in df.columns:
            df["Passenger Number"] = _normalize_phone(df["Passenger Number"])

        # Narrow to the columns we actually use for the minimal sheet
        df_clean = clean_file_without_headers(df)

        if df_clean is None or df_clean.empty:
            df_clean = pd.DataFrame(columns=[
                "First Name","Last Name","Passenger Number",
                "Member ID","Employee ID","Rider ID",
                "Pickup Time (Local)","Request Time (Local)","Pickup Time (UTC)","Request Time (UTC)",
                "Ride Status","Internal Note"
            ])

        return df_clean.fillna(""), df_raw.fillna("")


    except Exception as e:
        print("Error:", e)
        return None, None

def _normalize_phone(s: pd.Series) -> pd.Series:
    # to string, strip
    s = s.astype(object).fillna("").astype(str).str.strip()
    # drop trailing ".0" from float-looking values
    s = s.str.replace(r"\.0$", "", regex=True)
    # remove spaces, dashes, parentheses, etc. (keep digits and optional leading +)
    s = s.str.replace(r"[^\d+]", "", regex=True)
    # optional: strip leading US country code for 11-digit numbers
    s = s.str.replace(r"^\+?1(?=\d{10}$)", "", regex=True)
    return s

def build_daily_trip_sheet(
    df,
    include_refunds_bottom=True,
    internal_note_filter=None,   # str OR list[str]
):
    """
    ... (docstring unchanged)
    """
    import numpy as np
    import pandas as pd

    # Normalize filter input to a set of uppercase strings.
    # Accepts: None / "" / "All Internal Notes" (no filter) OR list of notes.
    def _normalize_filter(filter_in):
        if filter_in is None:
            return set()  # empty set = no filter
        if isinstance(filter_in, str):
            token = filter_in.strip().upper()
            if token in {"", "ALL", "ALL INTERNAL NOTES"}:
                return set()
            return {token}
        # list/iterable case
        try:
            vals = {str(x).strip().upper() for x in filter_in if str(x).strip()}
        except Exception:
            vals = set()
        return vals

    selected_set = _normalize_filter(internal_note_filter)

    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Pick up Date","Pick Up Time","First Name","Last Name","Rider Phone #",
            "Pick Up Address","Drop Off Address","Pickup to Drop of Miles","Transaction Amount",
            "Dispatcher Email","Internal Note","Transaction Type","Ride Status",
            "Trip Taken","Notes","Total Trips"
        ]).iloc[0:0]

    # -------- Optional Internal Note filter (applies to INPUT df) --------
    if selected_set and "Internal Note" in df.columns:
        s = df["Internal Note"].astype(str).str.strip().str.upper()
        df = df.loc[s.isin(selected_set)].copy()
        if df.empty:
            return pd.DataFrame(columns=[
                "Pick up Date","Pick Up Time","First Name","Last Name","Rider Phone #",
                "Pick Up Address","Drop Off Address","Pickup to Drop of Miles","Transaction Amount",
                "Dispatcher Email","Internal Note","Transaction Type","Ride Status",
                "Trip Taken","Notes","Total Trips"
            ]).iloc[0:0]

    # ---- remainder of function stays the same ----
    # (keep everything you already have below this point)

    # -------- Optional Internal Note filter (applies to INPUT df) --------
    if selected_set and "Internal Note" in df.columns:
        s = df["Internal Note"].astype(str).str.strip().str.upper()
        df = df.loc[s.isin(selected_set)].copy()
        if df.empty:
            return pd.DataFrame(columns=[
                "Pick up Date","Pick Up Time","First Name","Last Name","Rider Phone #",
                "Pick Up Address","Drop Off Address","Pickup to Drop of Miles","Transaction Amount",
                "Dispatcher Email","Internal Note","Transaction Type","Ride Status",
                "Trip Taken","Notes","Total Trips"
            ]).iloc[0:0]


    # ---- the rest of your original function (unchanged except for this header) ----

    # Buckets are defined globally above:
    # GROUP_A_NOTES = {"FCC","FCM","FCSH","FCSC","FCEX9","FCEX10"}
    # GROUP_B_NOTES = {"DTF","DTFCE"}

    df = df.copy()

    # --- choose time/date columns ---
    time_cols = ["Pickup Time (Local)", "Request Time (Local)", "Pickup Time (UTC)", "Request Time (UTC)"]
    time_col = next((c for c in time_cols if c in df.columns), None)
    df["_TIME_"] = df[time_col].fillna("").astype(str).str.strip() if time_col else ""
    date_col = next((c for c in ["Pickup Date (Local)", "Request Date (Local)"] if c in df.columns), None)
    df["_DATE_"] = df[date_col].fillna("").astype(str).str.strip() if date_col else ""

    # --- names ---
    first = df.get("First Name", pd.Series([""] * len(df))).fillna("").astype(str).str.strip()
    last  = df.get("Last Name",  pd.Series([""] * len(df))).fillna("").astype(str).str.strip()
    if (first.eq("").all() and last.eq("").all()) and ("Rider Name" in df.columns):
        name_series = df["Rider Name"].fillna("").astype(str)
        first_last = name_series.apply(_split_first_last)
        first = first_last.apply(lambda x: x[0])
        last  = first_last.apply(lambda x: x[1])

    # --- phone ---
    phone = df.get("Passenger Number", pd.Series([""] * len(df))).fillna("").astype(str).str.strip()
    phone = _normalize_phone(phone)

    # --- status/refund mapping (for Trip Taken) ---
    status_show = df.get("Ride Status", pd.Series([""] * len(df))).fillna("").astype(str).str.strip()
    amt_raw = df["Transaction Amount"] if "Transaction Amount" in df.columns else pd.Series([None] * len(df), index=df.index)
    amt_num = pd.to_numeric(pd.Series(amt_raw).astype(str).str.replace(r"[^\d\.\-]", "", regex=True), errors="coerce")
    refund_mask = amt_num.lt(0)
    status_show = status_show.mask(refund_mask, "REFUND")
    is_canceled = _is_canceled_series(status_show) | refund_mask

    # --- extras ---
    pickup_addr  = _col(df, "Pickup Address")
    dropoff_addr = _col(df, "Drop-off Address")
    if "Pickup to Drop-off Miles" in df.columns:
        miles = _col(df, "Pickup to Drop-off Miles")
    elif "Distance (miles)" in df.columns:
        miles = _col(df, "Distance (miles)")
    else:
        miles = pd.Series([""] * len(df), index=df.index)
    trans_amount = _col(df, "Transaction Amount")
    dispatcher   = _col(df, "Dispatcher Email")
    internal_src = "Internal Code" if "Internal Code" in df.columns else ("Internal Note" if "Internal Note" in df.columns else None)
    internal_vis = _col(df, internal_src) if internal_src else pd.Series([""] * len(df), index=df.index)
    trans_type   = _col(df, "Transaction Type")

    # --- assemble and HARD-FILTER canceled rows by Trip Taken flag ---
    trip_taken = np.select([refund_mask, is_canceled], ["-", "x"], default="✓")
    out = pd.DataFrame({
        "Pick up Date": df["_DATE_"],
        "Pick Up Time": df["_TIME_"],
        "First Name": first,
        "Last Name": last,
        "Rider Phone #": phone,
        "Pick Up Address": pickup_addr,
        "Drop Off Address": dropoff_addr,
        "Pickup to Drop of Miles": miles,
        "Transaction Amount": trans_amount,
        "Dispatcher Email": dispatcher,
        "Internal Note": internal_vis,
        "Transaction Type": trans_type,
        "Ride Status": status_show,
        "Trip Taken": trip_taken,
        "_is_canceled": is_canceled.values,
        "Notes": ''
    })

    # Normalize Trip Taken, filter canceled
    tt_norm = out["Trip Taken"].astype(str).str.normalize("NFKC").str.strip().str.lower()
    out = out.loc[~tt_norm.eq("x")].copy()

    # Helper flags/keys
    out["_is_refund"] = tt_norm.loc[out.index].eq("-").values
    out["_internal_sort"] = out["Internal Note"].astype(str).str.upper().str.strip()

    # Normalized sort keys
    out["_first_sort"] = _collapse_ws_series(out["First Name"]).str.upper()
    out["_last_sort"]  = _collapse_ws_series(out["Last Name"]).str.upper()
    out["_phone_sort"] = _collapse_ws_series(out["Rider Phone #"])

    # -------- Bucket by Internal Note --------
    def _matches_any_token(s: pd.Series, tokens: set[str]) -> pd.Series:
        if not tokens:
            return pd.Series(False, index=s.index)
        tokens_sorted = sorted(tokens, key=len, reverse=True)
        pat = r'(?<![A-Z0-9])(?:' + '|'.join(map(re.escape, tokens_sorted)) + r')'
        return s.str.contains(pat, regex=True, na=False)

    in_group_a = _matches_any_token(out["_internal_sort"], set(GROUP_A_NOTES))
    in_group_b = _matches_any_token(out["_internal_sort"], set(GROUP_B_NOTES))

    # Bucket: 0 = Group A (Fulton set), 1 = Group B (DTF/DTFCE), 2 = other/invalid
    out["_bucket"] = 2
    out.loc[in_group_a, "_bucket"] = 0
    out.loc[~in_group_a & in_group_b, "_bucket"] = 1

    # Visual sort
    sort_cols = ["_bucket", "_last_sort", "_first_sort", "_phone_sort", "_is_refund", "Pick Up Time"]
    out = out.sort_values(by=sort_cols, kind="stable", na_position="last").reset_index(drop=True)

    # Only A or B in main block; others go to "invalid" footer
    valid_mask = out["_bucket"].isin([0, 1])

    # Hide helper columns from export
    show_cols = [c for c in out.columns if c not in [
        "_is_canceled", "_is_refund", "_internal_sort",
        "_first_sort", "_last_sort", "_phone_sort",
        "_LNORM", "_FNORM", "_PNORM", "_bucket"
    ]]

    # refunds to show in footer (but KEEP them in main)
    refund_rows = out.loc[out["_is_refund"], show_cols].copy()

    # main block
    main_block = out.loc[valid_mask].copy()

    # Normalized keys for grouping/summing (space-collapsed, uppercased)
    if not main_block.empty:
        main_block["_LNORM"] = _collapse_ws_series(main_block["Last Name"]).str.upper()
        main_block["_FNORM"] = _collapse_ws_series(main_block["First Name"]).str.upper()
        main_block["_PNORM"] = _collapse_ws_series(main_block["Rider Phone #"])

    # ---- Totals per rider (EXCLUDING refunds) ----
    if not main_block.empty:
        main_block["Total Trips"] = ""
        grouped = main_block.groupby(["_LNORM","_FNORM","_PNORM"], sort=False, dropna=False)
        blocks = []
        show_cols_top = [c for c in show_cols if c != "Total Trips"] + ["Total Trips"]

        for i, (_, g) in enumerate(grouped):
            g = g.copy()
            non_refund_mask = ~g["_is_refund"]
            trip_count = int(non_refund_mask.sum())
            idx_place = g.loc[non_refund_mask].index.max() if non_refund_mask.any() else g.index.max()
            g.loc[idx_place, "Total Trips"] = trip_count
            blocks.append(g[show_cols_top])
            if i < len(grouped) - 1:
                blocks.append(pd.DataFrame([{c: "" for c in show_cols_top}]))

        final_df = pd.concat(blocks, ignore_index=True).fillna("")
    else:
        final_df = out[show_cols].iloc[0:0].copy()

    # footer: invalid internal notes (anything outside A/B)
    out_invalid = out.loc[~valid_mask, show_cols].copy()
    if not out_invalid.empty:
        spacer_above = pd.concat([pd.DataFrame([{c: "" for c in show_cols}]) for _ in range(10)], ignore_index=True)
        title_invalid = pd.DataFrame([{c: "" for c in show_cols}])
        title_invalid.iloc[0, 0] = "Internal note is not Forsyth ,Fulton, or an incorrect interal note was added"
        spacer_below = pd.concat([pd.DataFrame([{c: "" for c in show_cols}]) for _ in range(10)], ignore_index=True)
        final_df = pd.concat([final_df, spacer_above, title_invalid, out_invalid, spacer_below],
                             ignore_index=True).fillna("")

    # footer: refunds (duplicates on purpose)
    if include_refunds_bottom and not refund_rows.empty:
        final_df = pd.concat([final_df, pd.DataFrame([{c: "" for c in show_cols}])], ignore_index=True)
        title_refunds = pd.DataFrame([{c: "" for c in show_cols}])
        title_refunds.iloc[0, 0] = "Refunds"
        final_df = pd.concat([final_df, title_refunds, refund_rows], ignore_index=True).fillna("")

    return final_df

def sort_and_merge(file1_obj, file2_obj, internal_note_filter=None):  # str OR list[str]
    """
    Clean each file with the same pipeline, merge, optionally filter by Internal Note,
    then return the merged cleaned DataFrame.
    """
    import pandas as pd

    df1_clean, _ = clean_file(file1_obj)
    df2_clean, _ = clean_file(file2_obj)

    if df1_clean is None:
        df1_clean = pd.DataFrame()
    if df2_clean is None:
        df2_clean = pd.DataFrame()

    merged = pd.concat([df1_clean, df2_clean], ignore_index=True)

    # Normalize filter to a set (same helper logic as above, inlined here)
    def _normalize_filter(filter_in):
        if filter_in is None:
            return set()
        if isinstance(filter_in, str):
            token = filter_in.strip().upper()
            if token in {"", "ALL", "ALL INTERNAL NOTES"}:
                return set()
            return {token}
        try:
            vals = {str(x).strip().upper() for x in filter_in if str(x).strip()}
        except Exception:
            vals = set()
        return vals

    selected_set = _normalize_filter(internal_note_filter)

    # Optional Internal Note filter at the cleaned level
    if selected_set and "Internal Note" in merged.columns:
        s = merged["Internal Note"].astype(str).str.strip().str.upper()
        merged = merged.loc[s.isin(selected_set)].copy()

    # Tidy up (unchanged)
    for c in ["First Name", "Last Name", "Passenger Number"]:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).str.strip()

    sort_keys = [k for k in ["Last Name", "First Name", "Rider Phone #", "Pickup Time (Local)", "Request Time (Local)"] if k in merged.columns]
    if sort_keys:
        merged = merged.sort_values(by=sort_keys, kind="stable").reset_index(drop=True)

    return merged


# --- Streamlit UI ---
st.set_page_config(page_title="Daily Trip Counter", layout="centered")
st.title("📊 Daily Trip Counter")

# Controls
highlight_refunds = st.toggle(
    "Highlight refund rows (yellow)",
    value=False,
    help="When on, any row with a refund will be highlighted in yellow in the export."
)

include_refunds_bottom = st.checkbox(
    "Include Refunds at Bottom",
    value=True,
    help="When checked, refund rows are removed from the main table and listed in a footer section titled ‘Refunds’."
)

col1, col2 = st.columns(2)
with col1:
    uploaded_file_1 = st.file_uploader("Lyft File (.xlsx or .csv)", type=["xlsx", "csv"], key="file_1")
with col2:
    uploaded_file_2 = st.file_uploader("Uber File (.xlsx or .csv)", type=["xlsx", "csv"], key="file_2")

# 🔽 Filter controls BEFORE the button
all_notes = st.checkbox("Include all internal notes", value=True)
if all_notes:
    internal_note_filter = []  # empty list → no filter
else:
    internal_note_filter = st.multiselect(
        "Filter by Internal Note (choose one or many)",
        options=internal_note_values,
        default=[],  # or preselect a common subset if you like
        help="Leave empty to include none; use the checkbox above to include all."
    )

run = st.button("🧹 Gather daily count file")

if run:
    if not uploaded_file_1 and not uploaded_file_2:
        st.warning("Please upload at least one file.")
    else:
        # We may need both the cleaned (minimal) df and the raw df
        rider_only_df = None
        uber_detail_df = None
        out_filename = "daily_trips.xlsx"

        # CASE A: both files provided → merge cleaned + merge raw details
        if uploaded_file_1 and uploaded_file_2:
            # Build merged minimal using your helper, honoring the dropdown filter
            merged_clean = sort_and_merge(
                uploaded_file_1,
                uploaded_file_2,
                internal_note_filter=internal_note_filter
            )

            # Also build a merged RAW to feed the detail sheet (re-run clean_file to get raws)
            df1_clean, df1_raw = clean_file(uploaded_file_1)
            df2_clean, df2_raw = clean_file(uploaded_file_2)

            import pandas as pd
            # Avoid ambiguous truth-value on DataFrames
            df1_raw = df1_raw if df1_raw is not None else pd.DataFrame()
            df2_raw = df2_raw if df2_raw is not None else pd.DataFrame()
            merged_raw = pd.concat([df1_raw, df2_raw], ignore_index=True) if (not df1_raw.empty or not df2_raw.empty) else pd.DataFrame()

            # Build the rider-only daily sheet from the merged minimal (filtered)
            rider_only_df = build_daily_trip_sheet(
                merged_clean,
                include_refunds_bottom=include_refunds_bottom,
                internal_note_filter=internal_note_filter
            )
            # Build Uber detail from merged raw (unfiltered)
            uber_detail_df = build_uber_detail_sheet(merged_raw)

            out_filename = "daily_trips_merged.xlsx"

        # CASE B: only one file present → clean that one
        else:
            single = uploaded_file_1 if uploaded_file_1 else uploaded_file_2
            cleaned_df, raw_df = clean_file(single)

            if cleaned_df is None or cleaned_df.empty:
                st.error("❌ Could not clean this file or it has no valid rows.")
                st.stop()

            rider_only_df = build_daily_trip_sheet(
                cleaned_df,
                include_refunds_bottom=include_refunds_bottom,
                internal_note_filter=internal_note_filter   # <-- use the dropdown selection
            )
            uber_detail_df = build_uber_detail_sheet(raw_df)
            base = (single.name or "daily_trips").rsplit(".", 1)[0]
            out_filename = f"{base}_cleaned.xlsx"

        # ---- Display + Download ----
        if (rider_only_df is None or rider_only_df.empty) and (uber_detail_df is None or uber_detail_df.empty):
            st.info("No trips found.")
        else:
            st.success("✅ Sheets ready.")
            st.dataframe(rider_only_df.head(50) if rider_only_df is not None else None)

        # Build workbook buffer
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            if rider_only_df is not None:
                rider_only_df.to_excel(w, index=False, sheet_name="DailyTrips")
            if uber_detail_df is not None and not uber_detail_df.empty:
                # Uncomment if you want the Uber detail sheet included
                # uber_detail_df.to_excel(w, index=False, sheet_name="UberDetail")
                pass
        buf.seek(0)

        # --- Apply yellow highlight to refund rows if toggled ON ---
        if highlight_refunds and rider_only_df is not None and not rider_only_df.empty:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            # Determine which rows are refunds from the preview dataframe
            amt_num = pd.to_numeric(
                rider_only_df["Transaction Amount"].astype(str).str.replace(r"[^\d\.\-]", "", regex=True),
                errors="coerce"
            )
            refund_mask = rider_only_df["Trip Taken"].astype(str).eq("-") | amt_num.lt(0)

            if refund_mask.any():
                wb = load_workbook(buf)
                ws = wb["DailyTrips"]

                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Excel rows are 1-based and row 1 is header → add 2 to df index
                rows_to_color = (refund_mask[refund_mask].index + 2).tolist()
                max_col = ws.max_column

                for r in rows_to_color:
                    for c in range(1, max_col + 1):
                        ws[f"{get_column_letter(c)}{r}"].fill = fill

                # Save back into a fresh buffer
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)

        st.download_button(
            "📥 Download",
            buf,
            file_name=out_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
